import cv2
import numpy as np
import os
import openpyxl
import datetime

# Constants
SIZE = 4
HAAR_FILE = 'data/haarcascade_frontalface_default.xml'
DATASETS = 'database'
WIDTH = 130
HEIGHT = 100
THRESHOLD = 500
FILE_NAME = 'data.xlsx'

# Load the face recognition model
cv2.ocl.setUseOpenCL(False)
print('Training...')
(images, labels, names, id) = ([], [], {}, 0)
for (subdirs, dirs, files) in os.walk(DATASETS):
    for subdir in dirs:
        names[id] = subdir
        subjectpath = os.path.join(DATASETS, subdir)
        for filename in os.listdir(subjectpath):
            path = subjectpath + '/' + filename
            label = id
            images.append(cv2.imread(path, 0))
            labels.append(int(label))
        id += 1
(images, labels) = [np.array(lis) for lis in [images, labels]]
model = cv2.face.LBPHFaceRecognizer_create()
model.train(images, labels)

# Load the Excel workbook
if not os.path.isfile(FILE_NAME):
    wb = openpyxl.Workbook()
    wb.save(FILE_NAME)
wb = openpyxl.load_workbook(FILE_NAME)
ws = wb.active

# Read the student list from Excel sheet
student_list = []
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    student_list.append({
        'name': row[0],
        'phone': row[1],
        'present': row[2],
    })

# Start capturing video from webcam
face_cascade = cv2.CascadeClassifier(HAAR_FILE)
webcam = cv2.VideoCapture(0)

while True:
    # Read a frame from the webcam
    ret, im = webcam.read()
    if not ret:
        break
    # Convert the frame to grayscale and detect faces
    gray = cv2.cvtColor(im, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray, 1.3, 5)
    # Loop over all detected faces
    for (x, y, w, h) in faces:
        # Draw a rectangle around the face and resize it for recognition
        cv2.rectangle(im, (x, y), (x + w, y + h), (255, 0, 0), 2)
        face = gray[y:y + h, x:x + w]
        face_resize = cv2.resize(face, (WIDTH, HEIGHT))
        # Try to recognize the face
        prediction = model.predict(face_resize)
        cv2.rectangle(im, (x, y), (x + w, y + h), (0, 255, 0), 3)
        # If the prediction score is below the threshold, add the name to the Excel sheet
        if prediction[1] < THRESHOLD:
            name = names[prediction[0]]
            for student in student_list:
                if student['name'] == name:
                    # Check if the student is already marked as present
                    if not student['present']:
                        # Mark the student as present and record the time
                        student['present'] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        # Update the Excel sheet
                        updated = False
                        flag = 1
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                            flag += 1
                            if row[0] == student['name']:
                                ws.cell(flag, column=3, value=student['present'])
                                updated = True
                                wb.save(FILE_NAME)
                                break
            cv2.putText(im, name + ' - ' + str(prediction[1]), (x, y + h + 20), cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2, cv2.LINE_AA)
    # Display the image and check for key press
    cv2.imshow('Face Recognition', im)
    key = cv2.waitKey(10)
    if key == 27:
        break

# Release the webcam and save the Excel workbook
webcam.release()
cv2.destroyAllWindows()
wb.save(FILE_NAME)
print("Data saved to Excel sheet.")
print("Program terminated.")